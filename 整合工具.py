from openpyxl import Workbook,load_workbook, workbook
from openpyxl.utils import get_column_letter
from tools.fileManager import get_file_name
import os

def getValidColumnIndexByRow(worksheet, rowIndex):
    column = 1
    guideRow = list()
    while True:
        cell = get_column_letter(column) + str(rowIndex)
        value = worksheet[cell].value
        if column > 100:
            break
        elif value != None:
            guideRow.append(column)
        column = column + 1
    return guideRow    

def getRowValue(worksheet, rowIndex):
    # 補充
    GUIDE_ROW_VALID_COLUMN = getValidColumnIndexByRow(worksheet,1)
    NONE_IN_PYTHON = 'python_None'
    # DO
    rowValue = list()
    for col in GUIDE_ROW_VALID_COLUMN:
        cell = get_column_letter(col) + str(rowIndex)
        value = worksheet[cell].value
        rowValue.append(value)
    return rowValue     

def getSheetValue(worksheet):
    GUIDE_ROW_VALID_COLUMN = getValidColumnIndexByRow(worksheet,1)
    NULL_ROW = [None for col in GUIDE_ROW_VALID_COLUMN]
    # DO
    sheetValue = list()
    rowIndex = 1
    while True:
        rowValue = getRowValue(worksheet, rowIndex)
        if rowValue == NULL_ROW:
            break
        sheetValue.append(rowValue)
        rowIndex = rowIndex + 1 
    return sheetValue

def getBookValue():
    bookValue = list()
    for sheetIndex in range(len(workbook.worksheets)):
        worksheet = workbook.worksheets[sheetIndex]
        sheetValue = getSheetValue(worksheet)
        bookValue.append(sheetValue)
    return bookValue
        
def writeOnNewWorkbook():
    # 建立新的excel
    bookValue = getBookValue()
    newWorkbook = Workbook()
    newWorkSheet = newWorkbook.active
    # 渲染資料到新的excel工作表
    row = 1
    for sheetValue in bookValue:
        for rowValue in sheetValue:
            col = 1
            for cellValue in rowValue:
                cell = get_column_letter(col) + str(row)
                newWorkSheet[cell].value = cellValue
                # next
                col = col + 1 
            # next
            row = row + 1
    # 存檔
    try:
        newWorkbook.save('整合表.xlsx')
        print('--完成')
    except:
        print('--請關閉【整合表.xlsx】')
                    
                    
excel_file = get_file_name(__file__)
print('--找到excel檔案: ',excel_file)

workbook = load_workbook(excel_file,data_only=True)

writeOnNewWorkbook() 