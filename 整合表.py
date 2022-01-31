from openpyxl import Workbook,load_workbook, workbook
from openpyxl.utils import get_column_letter
import os

query_file_name = '整合表.xlsx'

def check_not_open_file(files):
    for file in files:
        if file == query_file_name:
            print('請關閉檔案:',query_file_name)
            os.system("pause") 
            return False
    return True
    
def get_file_name():
    # change directory to parent dir of this python file
    DIR_PATH = os.path.join(__file__, os.path.pardir)
    print(DIR_PATH)
    os.chdir(DIR_PATH)
    
    files = os.listdir()
    # check the query file isn't opened
    while True:
        files = os.listdir()
        if check_not_open_file('~$'+ query_file_name, files):
            break
    
    # find the excel file  
    for file in files:
        print(file)
        if '.xlsx' in file and file != query_file_name:
            return file
    return None


class Query:
    def __init__(self) -> None:
        excel_file = get_file_name()
        print('開啟檔案: ' + excel_file+ '...')
        OPEN_PATH = os.path.join(excel_file)
        
        self.wb = load_workbook(OPEN_PATH,data_only=True)
        self.wss = self.wb.worksheets
        self.ws_index = 0
        self.ws = self.wss[0]

    #private
    def goto_ws(self, index):
        try:
            self.ws_index = index
            self.ws = self.wss[self.ws_index]
        except:
            return None

    def get_column(self, find_value):
        GUIDE_ROW = 1
        column = 1
        while True:
            cell = get_column_letter(column) + str(GUIDE_ROW)
            value = self.ws[cell].value
            if column > 10:
                return None
            if find_value == value:
                return column
            column = column + 1

    def get_row_in_name(self, name_key):
        name_column = self.get_column('名稱')
        row = 2
        rows = list()
        while True:
            if row > 100:
                break
            cell = get_column_letter(name_column) + str(row)
            value = self.ws[cell].value
            if value == None:
                pass
            elif name_key in value:
                rows.append(row)
            row = row + 1
        return rows

    def valid_columns(self):
        GUIDE_ROW = 1
        column = 1
        columns = list()
        while True:
            cell = get_column_letter(column) + str(GUIDE_ROW)
            value = self.ws[cell].value
            if column > 100:
                break
            elif  value != None:
                columns.append(column)
            column = column + 1
        return columns

    def get_guide_row_data(self):
        cols = self.valid_columns()
        GUIDE_ROW = 1
        guide_row_data = list()
        for col in cols:
            cell = get_column_letter(col) + str(GUIDE_ROW)
            value = self.ws[cell].value
            guide_row_data.append(value)
        return guide_row_data

    def get_row_data(self, row):
        data = list()
        cols = self.valid_columns()
        for col in cols:
            cell = get_column_letter(col) + str(row)
            value = self.ws[cell].value
            data.append(value)
        return data

    def get_sheet_data(self, key):
        rows = self.get_row_in_name(key)
        sheet_data = list()
        sheet_data.append(self.get_guide_row_data())
        for row in rows:
            row_data = self.get_row_data(row)
            sheet_data.append(row_data)
        return sheet_data

    #public
    def get_book_data(self, key):
        book_data = list()
        for ws in self.wss:            
            sheet_data = self.get_sheet_data(key)
            if sheet_data.__len__() <= 1:
                self.goto_ws(self.ws_index + 1)
                continue
            for data in sheet_data:
                book_data.append(sheet_data)
            self.goto_ws(self.ws_index + 1)
        self.goto_ws(0)
        return book_data

    def write_on_new_wb(self, book_data):
        new_wb = Workbook()
        new_ws = new_wb.active
        row = 1
        for sheet_data in book_data:
            for row_data in sheet_data:
                col = 1
                for cell_data in row_data:
                    cell = get_column_letter(col) + str(row)
                    new_ws[cell].value = cell_data
                    col = col + 1 
                row = row + 1
        try:
            new_wb.save('篩選後資料.xlsx')
            print('完成')
        except:
            print('----請關閉或刪除「查詢檔案」')

    def start(self):
        key = input('--請輸入要查詢的產品名稱:')
        book_data = self.get_book_data(key)
        self.write_on_new_wb(book_data)